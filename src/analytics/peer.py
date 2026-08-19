"""Peer group analysis — Sprint 3.

Computes percentile ranks for 10 metrics within each of the 11 peer groups,
populates the `peer_percentiles` table, and exports a colour-coded
`output/peer_comparison.xlsx` (one sheet per group).
"""

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# (friendly name, financial_ratios column, invert_rank)
METRICS = [
    ("roe", "return_on_equity_pct", False),
    ("roce", "return_on_capital_employed_pct", False),
    ("net_profit_margin", "net_profit_margin_pct", False),
    ("debt_to_equity", "debt_to_equity", True),
    ("fcf", "free_cash_flow_cr", False),
    ("pat_cagr_5y", "pat_cagr_5yr", False),
    ("revenue_cagr_5y", "revenue_cagr_5yr", False),
    ("eps_cagr_5y", "eps_cagr_5yr", False),
    ("interest_coverage", "interest_coverage", False),
    ("asset_turnover", "asset_turnover", False),
]


class PeerAnalyzer:
    """Ranks companies within peer groups across 10 KPI metrics."""

    METRICS = METRICS

    def __init__(self, db_path: str = "db/nifty100.db"):
        self.db_path = Path(db_path)

    def _conn(self):
        return sqlite3.connect(self.db_path)

    def compute_percentiles(self, year: int | None = None) -> pd.DataFrame:
        conn = self._conn()
        try:
            groups = pd.read_sql(
                "SELECT company_id, peer_group_name, is_benchmark FROM peer_groups",
                conn,
            )
            if year is None:
                year = int(
                    pd.read_sql("SELECT MAX(year) AS y FROM financial_ratios", conn).iloc[0]["y"]
                )

            fr = pd.read_sql("SELECT * FROM financial_ratios WHERE year = ?", conn, params=[year])

            fr_renamed = fr.rename(columns={src: name for name, src, _ in METRICS})

            rows = []
            if not fr_renamed.empty:
                merged = fr_renamed.merge(groups, on="company_id", how="left")
                for group_name, grp in merged.groupby("peer_group_name"):
                    for name, _, invert in METRICS:
                        if name not in grp.columns:
                            continue
                        vals = grp[name].dropna()
                        if vals.empty or len(vals) < 2:
                            continue
                        pct = vals.rank(pct=True) * 100
                        if invert:
                            pct = 100 - pct
                        pct_map = dict(zip(grp["company_id"], pct))
                        for _, r in grp.iterrows():
                            rows.append(
                                {
                                    "company_id": int(r["company_id"]),
                                    "year": int(year),
                                    "metric": name,
                                    "value": r[name],
                                    "percentile_rank": pct_map.get(r["company_id"]),
                                    "peer_group": group_name,
                                }
                            )

            if rows:
                conn.execute("DELETE FROM peer_percentiles")
                pd.DataFrame(rows).to_sql("peer_percentiles", conn, if_exists="append", index=False)
                conn.commit()
        finally:
            conn.close()

        result = pd.DataFrame(rows)
        print(f"[Peer] Computed {len(result)} percentile rows for {year}")
        return result

    def _export_frame(self) -> pd.DataFrame:
        conn = self._conn()
        try:
            q = """SELECT pp.company_id, c.ticker, c.company_name, pp.peer_group,
                          pp.metric, pp.value, pp.percentile_rank
                   FROM peer_percentiles pp
                   JOIN companies c ON pp.company_id = c.company_id
                   ORDER BY pp.peer_group, c.ticker"""
            return pd.read_sql(q, conn)
        finally:
            conn.close()

    def export(self, path: str = "output/peer_comparison.xlsx") -> Path:
        data = self._export_frame()
        if data.empty:
            raise ValueError("No peer percentile data. Run compute_percentiles() first.")

        conn = self._conn()
        try:
            benchmarks = set(
                pd.read_sql("SELECT company_id FROM peer_groups WHERE is_benchmark = 1", conn)[
                    "company_id"
                ]
            )
        finally:
            conn.close()

        green = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        red = PatternFill("solid", fgColor="FFC7CE")
        gold = PatternFill("solid", fgColor="FFD966")
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        median_fill = PatternFill("solid", fgColor="D9E2F3")
        thin = Border(*[Side(style="thin")] * 4)

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        for group_name, grp in data.groupby("peer_group"):
            sheet = group_name[:31]
            ws = wb.create_sheet(title=sheet)

            pivot = grp.pivot_table(
                index=["company_id", "ticker", "company_name"],
                columns="metric",
                values="percentile_rank",
            )
            values_pivot = grp.pivot_table(
                index=["company_id", "ticker", "company_name"],
                columns="metric",
                values="value",
            )

            headers = ["company_id", "company_name"] + [f"{m} (pct)" for m, _, _ in METRICS]
            for cidx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=cidx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin

            row_num = 2
            for idx, pct_row in pivot.iterrows():
                cid = idx[0]
                is_bm = cid in benchmarks
                ws.cell(row=row_num, column=1, value=cid).border = thin
                ws.cell(row=row_num, column=2, value=idx[2]).border = thin
                for m_idx, (name, _, invert) in enumerate(METRICS, 3):
                    pct = pct_row.get(name)
                    cell = ws.cell(
                        row=row_num,
                        column=m_idx,
                        value=round(pct, 1) if pd.notna(pct) else None,
                    )
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin
                    if is_bm:
                        cell.fill = gold
                    elif pd.notna(pct):
                        if pct >= 75:
                            cell.fill = green
                        elif pct <= 25:
                            cell.fill = red
                        else:
                            cell.fill = yellow
                if is_bm:
                    ws.cell(row=row_num, column=1).fill = gold
                    ws.cell(row=row_num, column=2).fill = gold
                row_num += 1

            # Median row
            ws.cell(row=row_num, column=1, value="MEDIAN").fill = median_fill
            ws.cell(row=row_num, column=1).font = Font(bold=True, italic=True)
            ws.cell(row=row_num, column=2).fill = median_fill
            for m_idx, (name, _, _) in enumerate(METRICS, 3):
                vals = (
                    values_pivot[name].dropna() if name in values_pivot else pd.Series(dtype=float)
                )
                med = vals.median() if not vals.empty else None
                cell = ws.cell(
                    row=row_num,
                    column=m_idx,
                    value=round(med, 2) if pd.notna(med) else None,
                )
                cell.fill = median_fill
                cell.font = Font(bold=True, italic=True)
                cell.border = thin

            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 34
            for m_idx in range(3, len(headers) + 1):
                ws.column_dimensions[get_column_letter(m_idx)].width = 14

        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"[Peer] Exported {len(data.groupby('peer_group'))} sheets to {out}")
        return out


if __name__ == "__main__":
    a = PeerAnalyzer()
    a.compute_percentiles()
    a.export()
