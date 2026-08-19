"""Screener Engine — Sprint 3.

Loads latest-year financial ratios, applies threshold filters from
config/screener_config.yaml, computes the composite quality score
(0-100) and exports a colour-coded Excel workbook (one sheet per preset).

Composite quality score weights (per spec):
  35% Profitability (ROE 15 + ROCE 10 + NPM 10)
  30% Cash Quality  (FCF CAGR 15 + CFO/PAT 10 + FCF positive 5)
  20% Growth        (Revenue CAGR 10 + PAT CAGR 10)
  15% Leverage      (D/E 10 + ICR 5)
"""

import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.analytics.cagr import compute_cagr


class ScreenerEngine:
    """Applies preset filters to the latest-year financial ratios."""

    OUTPUT_COLS: list = [
        "company_id",
        "ticker",
        "company_name",
        "broad_sector",
        "roe",
        "roce",
        "net_profit_margin",
        "operating_profit_margin",
        "debt_to_equity",
        "interest_coverage",
        "free_cash_flow",
        "revenue_cagr_3y",
        "revenue_cagr_5y",
        "pat_cagr_5y",
        "eps_cagr_5y",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield",
        "market_cap",
        "net_profit",
        "sales",
        "composite_score",
    ]

    def __init__(self, config_path="config/screener_config.yaml", db_path="db/nifty100.db"):
        self.config_path = Path(config_path)
        self.db_path = Path(db_path)
        with open(self.config_path) as f:
            self.config = yaml.safe_load(f)
        self.presets = self.config["presets"]
        self._latest_year = None

    # ── Data loading ──────────────────────────────────────────────────

    def load_data(self) -> pd.DataFrame:
        conn = sqlite3.connect(self.db_path)
        try:
            latest = int(
                pd.read_sql("SELECT MAX(year) AS y FROM financial_ratios", conn).iloc[0]["y"]
            )
            self._latest_year = latest

            fr = pd.read_sql("SELECT * FROM financial_ratios WHERE year = ?", conn, params=[latest])
            comp = pd.read_sql(
                "SELECT company_id, ticker, company_name, broad_sector, sub_sector, "
                "market_cap_crore FROM companies",
                conn,
            )
            mc = pd.read_sql(
                "SELECT company_id, pe_ratio, pb_ratio, dividend_yield_pct FROM market_cap "
                "WHERE year = ?",
                conn,
                params=[latest],
            )
            pl = pd.read_sql(
                "SELECT company_id, sales, net_profit FROM profitandloss WHERE year = ?",
                conn,
                params=[latest],
            )

            df = (
                fr.merge(comp, on="company_id", how="left")
                .merge(mc, on="company_id", how="left")
                .merge(pl, on="company_id", how="left")
            )

            # D/E declining year-over-year
            prev = pd.read_sql(
                "SELECT company_id, debt_to_equity FROM financial_ratios WHERE year = ?",
                conn,
                params=[int(latest) - 1],
            )
            df = df.merge(
                prev.rename(columns={"debt_to_equity": "de_prev"}),
                on="company_id",
                how="left",
            )
            df["de_declining"] = df.apply(
                lambda r: (
                    pd.notna(r.get("debt_to_equity"))
                    and pd.notna(r.get("de_prev"))
                    and r["debt_to_equity"] < r["de_prev"]
                ),
                axis=1,
            )

            # FCF 5-year CAGR from cashflow
            cf = pd.read_sql(
                "SELECT company_id, year, operating_activity, investing_activity "
                "FROM cashflow ORDER BY company_id, year",
                conn,
            )
            df["fcf_cagr_5y"] = df["company_id"].map(self._fcf_cagr_map(cf))
        finally:
            conn.close()

        df = df.rename(
            columns={
                "return_on_equity_pct": "roe",
                "return_on_capital_employed_pct": "roce",
                "net_profit_margin_pct": "net_profit_margin",
                "operating_profit_margin_pct": "operating_profit_margin",
                "free_cash_flow_cr": "free_cash_flow",
                "revenue_cagr_3yr": "revenue_cagr_3y",
                "revenue_cagr_5yr": "revenue_cagr_5y",
                "pat_cagr_5yr": "pat_cagr_5y",
                "eps_cagr_5yr": "eps_cagr_5y",
                "dividend_payout_ratio_pct": "dividend_payout_ratio",
                "market_cap_crore": "market_cap",
                "dividend_yield_pct": "dividend_yield",
            }
        )
        self._df = df
        return df

    @staticmethod
    def _fcf_cagr_map(cf: pd.DataFrame) -> dict:
        if cf.empty:
            return {}
        cf = cf.copy()
        cf["fcf"] = cf["operating_activity"].fillna(0) + cf["investing_activity"].fillna(0)
        out = {}
        for cid, g in cf.groupby("company_id"):
            g = g.dropna(subset=["fcf"]).sort_values("year")
            if len(g) >= 6:
                val, _ = compute_cagr(g["fcf"].iloc[-6], g["fcf"].iloc[-1], 5)
                out[int(cid)] = val
            else:
                out[int(cid)] = None
        return out

    # ── Filtering ─────────────────────────────────────────────────────

    @staticmethod
    def _is_financial(sector) -> bool:
        return bool(sector) and "financial" in str(sector).lower()

    def apply_filters(self, df, preset_name) -> pd.DataFrame:
        preset = self.presets[preset_name]
        exclude = set(preset.get("exclude_financials_for", []))
        mask = pd.Series(True, index=df.index)

        for f in preset["filters"]:
            metric = f["metric"]
            min_v = f.get("min")
            max_v = f.get("max")
            op = f.get("operator")
            val = f.get("value")
            if metric not in df.columns:
                continue

            if op == "equals":
                fmask = df[metric].astype(str).str.lower() == str(val).lower()
            elif min_v is not None and max_v is not None:
                fmask = (df[metric] >= min_v) & (df[metric] <= max_v)
            elif min_v is not None:
                fmask = df[metric] >= min_v
            elif max_v is not None:
                fmask = df[metric] <= max_v
            else:
                continue

            # Debt Free companies always pass any ICR minimum threshold
            if metric == "interest_coverage" and min_v is not None and "icr_label" in df.columns:
                debt_free = df["icr_label"].eq("Debt Free")
                fmask = fmask | debt_free

            # D/E filters skip the Financials sector (structurally high leverage)
            if metric in exclude and "broad_sector" in df.columns:
                fmask = fmask | df["broad_sector"].apply(self._is_financial)

            mask = mask & fmask

        return df[mask]

    # ── Composite quality score ───────────────────────────────────────

    def composite_score(self, df) -> pd.DataFrame:
        result = df.copy()

        def rank_pct(series, invert=False):
            s = pd.to_numeric(series, errors="coerce")
            if s.notna().sum() < 2:
                return pd.Series(np.nan, index=s.index)
            lo, hi = s.quantile(0.10), s.quantile(0.90)
            w = s.clip(lo, hi)
            n = w.notna().sum()
            pct = ((w.rank() - 1) / max(n - 1, 1)) * 100
            return (100 - pct) if invert else pct

        roe = rank_pct(result.get("roe"))
        roce = rank_pct(result.get("roce"))
        npm = rank_pct(result.get("net_profit_margin"))
        fcf_cagr = rank_pct(result.get("fcf_cagr_5y"))
        cfo_pat = rank_pct(result.get("cfo_quality_score"))
        fcf_pos = (
            result.get("free_cash_flow", pd.Series(0, index=result.index)).fillna(0) > 0
        ) * 100.0
        rev_cagr = rank_pct(result.get("revenue_cagr_5y"))
        pat_cagr = rank_pct(result.get("pat_cagr_5y"))
        de = rank_pct(result.get("debt_to_equity"), invert=True)
        icr = rank_pct(result.get("interest_coverage"))

        profitability = (roe * 15 + roce * 10 + npm * 10) / 35
        cash_quality = (fcf_cagr * 15 + cfo_pat * 10 + fcf_pos * 5) / 30
        growth = (rev_cagr * 10 + pat_cagr * 10) / 20
        leverage = (de * 10 + icr * 5) / 15

        composite = profitability * 0.35 + cash_quality * 0.30 + growth * 0.20 + leverage * 0.15
        result["composite_score"] = composite.clip(0, 100)

        if "broad_sector" in result.columns:
            result["composite_score_sector"] = (
                result.groupby("broad_sector")["composite_score"].rank(pct=True) * 100
            )
        return result

    # ── Screening + export ────────────────────────────────────────────

    def screen(self, preset_name) -> pd.DataFrame:
        if preset_name not in self.presets:
            raise ValueError(
                f"Unknown preset '{preset_name}'. Available: {list(self.presets.keys())}"
            )
        df = self.load_data()
        filtered = self.apply_filters(df, preset_name)
        if filtered.empty:
            return filtered
        scored = self.composite_score(filtered)
        scored = scored.sort_values("composite_score", ascending=False)
        return scored[
            (
                self.OUTPUT_COLS + ["composite_score_sector"]
                if "composite_score_sector" in scored.columns
                else self.OUTPUT_COLS
            )
        ]

    def _write_sheet(self, ws, df, preset):
        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Border(*[Side(style="thin")] * 4)

        cols = list(df.columns)
        for cidx, cname in enumerate(cols, 1):
            cell = ws.cell(row=1, column=cidx, value=cname)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin

        filter_rules = preset.get("filters", [])
        rule_map = {r["metric"]: r for r in filter_rules if "metric" in r}

        for ridx, (_, row) in enumerate(df.iterrows(), 2):
            for cidx, cname in enumerate(cols, 1):
                val = row[cname]
                if isinstance(val, float) and (pd.isna(val) or np.isnan(val)):
                    val = None
                cell = ws.cell(
                    row=ridx,
                    column=cidx,
                    value=round(val, 2) if isinstance(val, float) else val,
                )
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin

                rule = rule_map.get(cname)
                if rule and val is not None and not isinstance(val, bool):
                    min_v, max_v = rule.get("min"), rule.get("max")
                    try:
                        if min_v is not None and val < min_v or max_v is not None and val > max_v:
                            cell.fill = red
                        else:
                            cell.fill = green
                    except (TypeError, ValueError):
                        pass

            for cidx, cname in enumerate(cols, 1):
                if cname == "composite_score":
                    ws.column_dimensions[get_column_letter(cidx)].width = 14

    def run_all(self, output_path="output/screener_output.xlsx") -> Path:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        results = {}
        for preset_name in self.presets:
            sheet_name = preset_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            try:
                res = self.screen(preset_name)
                results[preset_name] = res
                self._write_sheet(ws, res, self.presets[preset_name])
            except Exception as e:  # pragma: no cover
                ws["A1"] = f"Error: {e}"
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"[Screener] Exported to {out}")
        for name, res in results.items():
            print(f"  {name}: {len(res)} companies")
        return out


if __name__ == "__main__":
    ScreenerEngine().run_all()
